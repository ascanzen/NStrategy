from __future__ import annotations

import json
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

import pandas as pd
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(__file__).resolve().parents[1]
N_PATTERN_DIR = Path(os.environ.get("N_PATTERN_DIR", ROOT_DIR / "outputs" / "n_pattern")).resolve()
TRANSITION_FILE = N_PATTERN_DIR / "n_transition_recent.csv"
SUMMARY_FILE = N_PATTERN_DIR / "n_transition_recent_summary.json"
GLOBAL_SUMMARY_FILE = N_PATTERN_DIR / "n_summary.json"
ALIAS_FILES = [
    Path(os.environ.get("STOCK_ALIAS_FILE", "")) if os.environ.get("STOCK_ALIAS_FILE") else None,
    ROOT_DIR / "data" / "stock_aliases.csv",
    ROOT_DIR / "data" / "stock_aliases.json",
]

app = FastAPI(title="N Pattern Browser", version="1.0.0")
SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:8080",
        "http://127.0.0.1:8080",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _read_csv(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size <= 1:
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except pd.errors.EmptyDataError:
        return pd.DataFrame()


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def _normalize_code(value: Any) -> str:
    text = str(value or "").upper().strip()
    return text.replace(".", "").replace("-", "").replace("_", "")


def _safe_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value)


def _now_shanghai() -> datetime:
    return datetime.now(SHANGHAI_TZ)


def _to_pinyin(value: Any) -> str:
    text = _safe_text(value)
    if not text:
        return ""
    try:
        from pypinyin import lazy_pinyin  # type: ignore
    except Exception:
        return ""
    return " ".join(lazy_pinyin(text)).lower()


def _pinyin_abbr(value: Any) -> str:
    text = _safe_text(value)
    if not text:
        return ""
    return "".join(part[:1] for part in text.split() if part).lower()


def _load_aliases() -> dict[str, dict[str, str]]:
    aliases: dict[str, dict[str, str]] = {}
    for path in ALIAS_FILES:
        if path is None or not path.exists():
            continue
        try:
            if path.suffix.lower() == ".json":
                rows = json.loads(path.read_text(encoding="utf-8"))
                df = pd.DataFrame(rows if isinstance(rows, list) else rows.get("rows", []))
            else:
                df = pd.read_csv(path)
        except Exception:
            continue
        if df.empty:
            continue
        for row in df.to_dict("records"):
            keys = {
                _normalize_code(row.get("ts_code")),
                _normalize_code(row.get("instrument")),
                _normalize_code(row.get("code")),
                _normalize_code(row.get("symbol")),
            }
            payload = {
                "stock_name": _safe_text(row.get("name") or row.get("stock_name")),
                "stock_pinyin": _safe_text(row.get("pinyin") or row.get("stock_pinyin")),
                "stock_pinyin_abbr": _safe_text(row.get("abbr") or row.get("stock_pinyin_abbr")),
            }
            if payload["stock_name"] and not payload["stock_pinyin"]:
                payload["stock_pinyin"] = _to_pinyin(payload["stock_name"])
            if payload["stock_pinyin"] and not payload["stock_pinyin_abbr"]:
                payload["stock_pinyin_abbr"] = _pinyin_abbr(payload["stock_pinyin"])
            for key in keys:
                if key:
                    aliases[key] = payload
    return aliases


def _transition_data() -> pd.DataFrame:
    df = _read_csv(TRANSITION_FILE)
    if df.empty:
        return df

    for column in ["trade_date", "datetime", "ts_code", "instrument", "n_transition", "n_signal_name", "n_prev_signal_name", "n_svg_path"]:
        if column not in df.columns:
            df[column] = ""

    aliases = _load_aliases()
    if aliases:
        names: list[str] = []
        pinyin: list[str] = []
        abbr: list[str] = []
        for row in df.to_dict("records"):
            key = _normalize_code(row.get("ts_code")) or _normalize_code(row.get("instrument"))
            alias = aliases.get(key, {})
            names.append(alias.get("stock_name", ""))
            pinyin.append(alias.get("stock_pinyin", ""))
            abbr.append(alias.get("stock_pinyin_abbr", ""))
        df["stock_name"] = names
        df["stock_pinyin"] = pinyin
        df["stock_pinyin_abbr"] = abbr

    df["trade_date"] = df["trade_date"].astype(str)
    return df


def _recent_dates(df: pd.DataFrame, days: int) -> list[str]:
    if df.empty or "trade_date" not in df.columns:
        return []
    dates = sorted({str(value) for value in df["trade_date"].dropna() if str(value)})
    return dates[-days:]


def _matches_query(row: pd.Series, query: str) -> bool:
    if not query:
        return True
    q = query.strip().lower()
    q_code = _normalize_code(q)
    fields = [
        row.get("ts_code"),
        row.get("instrument"),
        row.get("stock_name"),
        row.get("stock_pinyin"),
        row.get("stock_pinyin_abbr"),
    ]
    for value in fields:
        text = _safe_text(value)
        if not text:
            continue
        lower_text = text.lower()
        compact_text = lower_text.replace(" ", "")
        if q in lower_text or q in compact_text or q_code in _normalize_code(text):
            return True
    return False


def _apply_filters(
    df: pd.DataFrame,
    *,
    days: int,
    trade_date: str | None,
    transition: str | None,
    query: str | None,
) -> pd.DataFrame:
    if df.empty:
        return df

    filtered = df.copy()
    if trade_date:
        filtered = filtered[filtered["trade_date"].astype(str) == trade_date]
    else:
        dates = _recent_dates(filtered, days)
        filtered = filtered[filtered["trade_date"].astype(str).isin(dates)]

    if transition and transition != "all":
        filtered = filtered[filtered["n_transition"].astype(str) == transition]

    if query:
        filtered = filtered[filtered.apply(lambda row: _matches_query(row, query), axis=1)]

    sort_columns = [column for column in ["trade_date", "n_transition", "ts_code"] if column in filtered.columns]
    if sort_columns:
        filtered = filtered.sort_values(sort_columns, ascending=[False] + [True] * (len(sort_columns) - 1))
    return filtered


def _record_payload(row: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for key, value in row.items():
        if pd.isna(value):
            payload[key] = None
        elif hasattr(value, "item"):
            payload[key] = value.item()
        else:
            payload[key] = value

    svg_path = _safe_text(payload.get("n_svg_path"))
    payload["svg_url"] = f"/api/svg/{quote(svg_path)}" if svg_path else None
    payload["display_name"] = payload.get("stock_name") or payload.get("instrument") or payload.get("ts_code")
    return payload


def _filtered_transitions(
    days: int,
    trade_date: str | None,
    transition: str | None,
    query: str | None,
) -> pd.DataFrame:
    days = max(1, min(days, 30))
    df = _transition_data()
    return _apply_filters(df, days=days, trade_date=trade_date, transition=transition, query=query)


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {
        "ok": True,
        "n_pattern_dir": str(N_PATTERN_DIR),
        "transition_file_exists": TRANSITION_FILE.exists(),
        "generated_at": _now_shanghai().isoformat(timespec="seconds"),
    }


@app.get("/api/summary")
def summary() -> dict[str, Any]:
    df = _transition_data()
    dates = _recent_dates(df, 30)
    transition_counts = {}
    signal_counts = {}
    if not df.empty:
        transition_counts = df["n_transition"].fillna("unknown").value_counts().to_dict()
        signal_counts = df["n_signal_name"].fillna("unknown").value_counts().to_dict()
    return {
        "n_pattern_dir": str(N_PATTERN_DIR),
        "transition_summary": _read_json(SUMMARY_FILE),
        "global_summary": _read_json(GLOBAL_SUMMARY_FILE),
        "available_dates": dates,
        "row_count": int(len(df)),
        "transition_counts": transition_counts,
        "signal_counts": signal_counts,
    }


@app.get("/api/transitions")
def transitions(
    days: int = Query(3, ge=1, le=30),
    trade_date: str | None = Query(None),
    transition: str = Query("all"),
    q: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
) -> dict[str, Any]:
    filtered = _filtered_transitions(days, trade_date, transition, q)
    page = filtered.iloc[offset : offset + limit]
    records = [_record_payload(row) for row in page.to_dict("records")]
    return {
        "total": int(len(filtered)),
        "limit": limit,
        "offset": offset,
        "items": records,
    }


@app.get("/api/svg/{svg_path:path}")
def svg(svg_path: str) -> FileResponse:
    target = (N_PATTERN_DIR / svg_path).resolve()
    try:
        target.relative_to(N_PATTERN_DIR)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="SVG not found") from exc
    if not target.exists() or target.suffix.lower() != ".svg":
        raise HTTPException(status_code=404, detail="SVG not found")
    return FileResponse(target, media_type="image/svg+xml")


@app.get("/api/export.xlsx")
def export_xlsx(
    days: int = Query(3, ge=1, le=30),
    trade_date: str | None = Query(None),
    transition: str = Query("all"),
    q: str | None = Query(None),
) -> StreamingResponse:
    filtered = _filtered_transitions(days, trade_date, transition, q)
    wb = Workbook()
    ws = wb.active
    ws.title = "N Pattern"

    columns = [
        ("trade_date", "日期"),
        ("ts_code", "股票代码"),
        ("instrument", "标的"),
        ("stock_name", "名称"),
        ("n_transition", "N型切换"),
        ("n_signal_name", "当前N型"),
        ("n_prev_signal_name", "前一N型"),
        ("n_prev_trade_date", "前一日期"),
        ("close", "收盘价"),
        ("vol", "成交量"),
        ("n_pivot_count", "拐点数"),
        ("n_window", "窗口"),
        ("n_threshold", "阈值"),
        ("n_svg_path", "SVG路径"),
    ]
    ws.append([label for _, label in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")

    for row in filtered.to_dict("records"):
        ws.append([None if pd.isna(row.get(key)) else row.get(key) for key, _ in columns])

    for index, (_, label) in enumerate(columns, start=1):
        width = max(len(label) + 4, 12)
        if label in {"SVG路径", "N型切换"}:
            width = 42
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"n_pattern_transitions_{_now_shanghai().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
